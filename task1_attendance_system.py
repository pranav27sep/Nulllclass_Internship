import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import cv2
import os
import csv
import json
import time
import threading
import datetime
import numpy as np
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import torch
import torch.nn as nn
import torchvision.transforms as transforms
import torchvision.models as models
import pickle

WINDOW_START = datetime.time(9, 30)   
WINDOW_END   = datetime.time(10, 0)
ATTENDANCE_FILE = "attendance_records.xlsx"
STUDENT_DB_FILE = "student_database.pkl"
EMBEDDINGS_DIR  = "student_faces"
DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

EMOTIONS = ["Angry", "Disgust", "Fear", "Happy", "Sad", "Surprise", "Neutral"]

class FaceEmbeddingModel(nn.Module):
    """ResNet18-based face embedding network (128-d output)."""
    def __init__(self):
        super().__init__()
        base = models.resnet18(weights=models.ResNet18_Weights.DEFAULT)
        self.features = nn.Sequential(*list(base.children())[:-1])
        self.embed = nn.Linear(512, 128)

    def forward(self, x):
        x = self.features(x)
        x = x.flatten(1)
        x = self.embed(x)
        return nn.functional.normalize(x, dim=1)

# ── PyTorch Emotion Model ─────────────────────────────────────
class EmotionModel(nn.Module):
    """MobileNetV2-based emotion classifier (7 classes)."""
    def __init__(self):
        super().__init__()
        base = models.mobilenet_v2(weights=models.MobileNet_V2_Weights.DEFAULT)
        base.features[0][0] = nn.Conv2d(1, 32, 3, stride=2, padding=1, bias=False)
        base.classifier = nn.Sequential(
            nn.Dropout(0.2),
            nn.Linear(base.last_channel, 7)
        )
        self.net = base

    def forward(self, x):
        return self.net(x)

# ── Student Database ──────────────────────────────────────────
class StudentDatabase:
    def __init__(self, path=STUDENT_DB_FILE):
        self.path = path
        self.students: dict[str, dict] = {}  # id -> {name, embedding}
        self.load()

    def load(self):
        if os.path.exists(self.path):
            with open(self.path, "rb") as f:
                self.students = pickle.load(f)

    def save(self):
        with open(self.path, "wb") as f:
            pickle.dump(self.students, f)

    def add_student(self, student_id: str, name: str, embedding: np.ndarray):
        self.students[student_id] = {"name": name, "embedding": embedding}
        self.save()

    def find_match(self, embedding: np.ndarray, threshold=0.6):
        """Return (student_id, name, distance) or None."""
        best, best_dist = None, float("inf")
        for sid, info in self.students.items():
            db_emb = info["embedding"]
            dist = float(np.linalg.norm(embedding - db_emb))
            if dist < best_dist:
                best_dist = dist
                best = (sid, info["name"], dist)
        if best and best_dist < threshold:
            return best
        return None

# ── Attendance Recorder ───────────────────────────────────────
class AttendanceRecorder:
    def __init__(self, file=ATTENDANCE_FILE):
        self.file = file
        self.today = datetime.date.today().isoformat()
        self.records: dict[str, dict] = {}   # student_id -> record

    def mark(self, student_id, name, emotion, confidence):
        if student_id not in self.records:
            self.records[student_id] = {
                "student_id": student_id,
                "name": name,
                "status": "Present",
                "emotion": emotion,
                "confidence": round(confidence, 3),
                "time": datetime.datetime.now().strftime("%H:%M:%S"),
                "date": self.today,
            }

    def mark_absents(self, all_students: dict):
        for sid, info in all_students.items():
            if sid not in self.records:
                self.records[sid] = {
                    "student_id": sid,
                    "name": info["name"],
                    "status": "Absent",
                    "emotion": "N/A",
                    "confidence": 0.0,
                    "time": "N/A",
                    "date": self.today,
                }

    def save_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Header styling
        header_fill = PatternFill("solid", fgColor="2E4057")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        headers = ["Student ID", "Name", "Status", "Emotion", "Confidence", "Time", "Date"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        present_fill = PatternFill("solid", fgColor="C8F7C5")
        absent_fill  = PatternFill("solid", fgColor="FADBD8")
        for row_idx, rec in enumerate(self.records.values(), 2):
            row = [rec["student_id"], rec["name"], rec["status"],
                   rec["emotion"], rec["confidence"], rec["time"], rec["date"]]
            fill = present_fill if rec["status"] == "Present" else absent_fill
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        wb.save(self.file)
        return self.file

    def save_csv(self):
        csv_file = self.file.replace(".xlsx", ".csv")
        with open(csv_file, "w", newline="") as f:
            fields = ["student_id", "name", "status", "emotion", "confidence", "time", "date"]
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            w.writerows(self.records.values())
        return csv_file

# ── GUI Application ───────────────────────────────────────────
class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🎓 Smart Attendance System — PyTorch")
        self.root.geometry("1200x750")
        self.root.configure(bg="#1A1A2E")
        self.root.resizable(True, True)

        # Models
        self.embed_model = FaceEmbeddingModel().to(DEVICE).eval()
        self.emo_model   = EmotionModel().to(DEVICE).eval()
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

        # Data
        self.db = StudentDatabase()
        self.recorder = AttendanceRecorder()
        self.cap = None
        self.running = False
        self.frame_lock = threading.Lock()
        self.current_frame = None

        # Transforms
        self.embed_tf = transforms.Compose([
            transforms.Resize((112, 112)),
            transforms.ToTensor(),
            transforms.Normalize([0.5]*3, [0.5]*3),
        ])
        self.emo_tf = transforms.Compose([
            transforms.Resize((48, 48)),
            transforms.Grayscale(),
            transforms.ToTensor(),
            transforms.Normalize([0.5], [0.5]),
        ])

        self._build_ui()
        self._check_time_window()

    # ── UI ────────────────────────────────────────────────────
    def _build_ui(self):
        # Top bar
        top = tk.Frame(self.root, bg="#16213E", height=60)
        top.pack(fill=tk.X)
        tk.Label(top, text="🎓 SMART ATTENDANCE SYSTEM", font=("Courier", 18, "bold"),
                 fg="#E94560", bg="#16213E").pack(side=tk.LEFT, padx=20, pady=10)
        self.time_lbl = tk.Label(top, text="", font=("Courier", 13), fg="#A8DADC", bg="#16213E")
        self.time_lbl.pack(side=tk.RIGHT, padx=20)
        self._tick()

        # Main panes
        pane = tk.PanedWindow(self.root, orient=tk.HORIZONTAL, bg="#1A1A2E", sashwidth=4)
        pane.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)

        # Left: camera feed
        left = tk.Frame(pane, bg="#16213E", relief=tk.FLAT)
        pane.add(left, width=640)
        tk.Label(left, text="📷 LIVE CAMERA FEED", font=("Courier", 11, "bold"),
                 fg="#E94560", bg="#16213E").pack(pady=(8, 2))
        self.video_lbl = tk.Label(left, bg="#0F3460", width=620, height=440)
        self.video_lbl.pack(padx=6, pady=4)

        # Camera controls
        ctrl = tk.Frame(left, bg="#16213E")
        ctrl.pack(fill=tk.X, pady=4, padx=6)
        self._btn(ctrl, "▶ Start Camera", self.start_camera, "#E94560").pack(side=tk.LEFT, padx=4)
        self._btn(ctrl, "⏹ Stop", self.stop_camera, "#555").pack(side=tk.LEFT, padx=4)
        self._btn(ctrl, "📸 Capture & Mark", self.capture_mark, "#06D6A0").pack(side=tk.LEFT, padx=4)

        # Register student
        reg = tk.LabelFrame(left, text=" ➕ Register New Student ", bg="#16213E",
                            fg="#A8DADC", font=("Courier", 10, "bold"), bd=1, relief=tk.GROOVE)
        reg.pack(fill=tk.X, padx=6, pady=4)
        r1 = tk.Frame(reg, bg="#16213E"); r1.pack(fill=tk.X, pady=4)
        tk.Label(r1, text="ID:", bg="#16213E", fg="#A8DADC", width=8).pack(side=tk.LEFT)
        self.reg_id = tk.Entry(r1, width=12, bg="#0F3460", fg="white", insertbackground="white")
        self.reg_id.pack(side=tk.LEFT, padx=4)
        tk.Label(r1, text="Name:", bg="#16213E", fg="#A8DADC").pack(side=tk.LEFT)
        self.reg_name = tk.Entry(r1, width=16, bg="#0F3460", fg="white", insertbackground="white")
        self.reg_name.pack(side=tk.LEFT, padx=4)
        self._btn(r1, "Register", self.register_student, "#FFC300").pack(side=tk.LEFT, padx=4)

        # Right: status & log
        right = tk.Frame(pane, bg="#16213E")
        pane.add(right)

        # Time window status
        self.window_lbl = tk.Label(right, text="", font=("Courier", 11, "bold"),
                                   bg="#16213E", pady=6)
        self.window_lbl.pack(fill=tk.X, padx=10)

        # Stats row
        stats = tk.Frame(right, bg="#16213E")
        stats.pack(fill=tk.X, padx=10, pady=4)
        self.stat_present = self._stat_card(stats, "Present", "#06D6A0")
        self.stat_absent  = self._stat_card(stats, "Absent",  "#E94560")
        self.stat_total   = self._stat_card(stats, "Students", "#A8DADC")

        # Attendance table
        tk.Label(right, text="📋 ATTENDANCE LOG", font=("Courier", 11, "bold"),
                 fg="#E94560", bg="#16213E").pack(pady=(8,2))
        cols = ("ID", "Name", "Status", "Emotion", "Time")
        self.tree = ttk.Treeview(right, columns=cols, show="headings", height=14)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#0F3460", foreground="white",
                        fieldbackground="#0F3460", rowheight=26, font=("Courier", 10))
        style.configure("Treeview.Heading", background="#E94560", foreground="white",
                        font=("Courier", 10, "bold"))
        style.map("Treeview", background=[("selected", "#2D6A4F")])
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=90, anchor=tk.CENTER)
        self.tree.tag_configure("present", background="#1B4332")
        self.tree.tag_configure("absent",  background="#641220")
        sb = ttk.Scrollbar(right, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10,0), pady=4)
        sb.pack(side=tk.LEFT, fill=tk.Y, pady=4)

        # Bottom controls
        bot = tk.Frame(self.root, bg="#16213E")
        bot.pack(fill=tk.X, padx=10, pady=6)
        self._btn(bot, "💾 Save Excel", self.save_excel, "#06D6A0").pack(side=tk.LEFT, padx=4)
        self._btn(bot, "📄 Save CSV",   self.save_csv,   "#A8DADC").pack(side=tk.LEFT, padx=4)
        self._btn(bot, "🔄 Mark Absents", self.mark_all_absent, "#FFC300").pack(side=tk.LEFT, padx=4)
        self._btn(bot, "🗑 Clear Log",  self.clear_log,  "#555").pack(side=tk.LEFT, padx=4)
        self.status_bar = tk.Label(bot, text="Ready", fg="#A8DADC", bg="#16213E",
                                   font=("Courier", 10))
        self.status_bar.pack(side=tk.RIGHT, padx=10)

    def _btn(self, parent, text, cmd, color="#E94560"):
        b = tk.Button(parent, text=text, command=cmd, bg=color, fg="white",
                      font=("Courier", 9, "bold"), relief=tk.FLAT, padx=8, pady=4,
                      activebackground="#555", cursor="hand2")
        return b

    def _stat_card(self, parent, label, color):
        f = tk.Frame(parent, bg="#0F3460", padx=12, pady=6, relief=tk.GROOVE, bd=1)
        f.pack(side=tk.LEFT, padx=6, expand=True, fill=tk.X)
        num = tk.Label(f, text="0", font=("Courier", 20, "bold"), fg=color, bg="#0F3460")
        num.pack()
        tk.Label(f, text=label, font=("Courier", 9), fg="#A8DADC", bg="#0F3460").pack()
        return num

    # ── Clock ─────────────────────────────────────────────────
    def _tick(self):
        now = datetime.datetime.now().strftime("%A  %H:%M:%S")
        self.time_lbl.config(text=now)
        self.root.after(1000, self._tick)

    def _check_time_window(self):
        now_t = datetime.datetime.now().time()
        if WINDOW_START <= now_t <= WINDOW_END:
            self.window_lbl.config(text=f"✅ Attendance Window OPEN  [{WINDOW_START}–{WINDOW_END}]",
                                   fg="#06D6A0", bg="#16213E")
            self.in_window = True
        else:
            self.window_lbl.config(
                text=f"⛔ Outside Window  (Open: {WINDOW_START}–{WINDOW_END})",
                fg="#E94560", bg="#16213E")
            self.in_window = False
        self.root.after(5000, self._check_time_window)

    # ── Camera ────────────────────────────────────────────────
    def start_camera(self):
        if self.running:
            return
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            # Demo mode: generate synthetic frames
            self.status_bar.config(text="📷 No camera – Demo mode active")
            self.running = True
            threading.Thread(target=self._demo_loop, daemon=True).start()
            return
        self.running = True
        threading.Thread(target=self._camera_loop, daemon=True).start()
        self.status_bar.config(text="📷 Camera running…")

    def stop_camera(self):
        self.running = False
        if self.cap:
            self.cap.release()
            self.cap = None
        self.status_bar.config(text="Camera stopped")

    def _camera_loop(self):
        while self.running:
            ret, frame = self.cap.read()
            if not ret:
                break
            frame = self._annotate(frame)
            with self.frame_lock:
                self.current_frame = frame.copy()
            self._show_frame(frame)
            time.sleep(0.03)

    def _demo_loop(self):
        """Synthetic frame for demo when no camera is connected."""
        idx = 0
        while self.running:
            frame = np.zeros((480, 640, 3), dtype=np.uint8)
            frame[:] = (15, 33, 62)
            cv2.putText(frame, "DEMO MODE - No Camera", (120, 200),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (232, 69, 96), 2)
            cv2.putText(frame, "Register students & use 'Capture & Mark'", (50, 260),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (168, 218, 220), 1)
            idx += 1
            with self.frame_lock:
                self.current_frame = frame.copy()
            self._show_frame(frame)
            time.sleep(0.05)

    def _annotate(self, frame):
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(gray, 1.1, 4, minSize=(60, 60))
        for (x, y, w, h) in faces:
            face_img = frame[y:y+h, x:x+w]
            match = self._identify(face_img)
            emotion = self._detect_emotion(face_img)
            if match:
                sid, name, dist = match
                label = f"{name} ({emotion})"
                color = (0, 255, 100)
            else:
                label = f"Unknown ({emotion})"
                color = (0, 100, 255)
            cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
            cv2.rectangle(frame, (x, y-28), (x+w, y), color, -1)
            cv2.putText(frame, label, (x+4, y-8),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 0, 0), 1)
        return frame

    def _show_frame(self, frame):
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(rgb).resize((620, 440))
        photo = ImageTk.PhotoImage(img)
        self.video_lbl.configure(image=photo)
        self.video_lbl.image = photo

    # ── Identification ────────────────────────────────────────
    def _get_embedding(self, face_bgr):
        face_rgb = cv2.cvtColor(face_bgr, cv2.COLOR_BGR2RGB)
        pil = Image.fromarray(face_rgb)
        t = self.embed_tf(pil).unsqueeze(0).to(DEVICE)
        with torch.no_grad():
            emb = self.embed_model(t).cpu().numpy()[0]
        return emb

    def _identify(self, face_bgr):
        if face_bgr.size == 0:
            return None
        try:
            emb = self._get_embedding(face_bgr)
            return self.db.find_match(emb)
        except Exception:
            return None

    def _detect_emotion(self, face_bgr):
        try:
            gray = cv2.cvtColor(face_bgr, cv2.COLOR_BGR2GRAY)
            pil = Image.fromarray(gray)
            t = self.emo_tf(pil).unsqueeze(0).to(DEVICE)
            with torch.no_grad():
                out = self.emo_model(t)
                idx = out.argmax(1).item()
            return EMOTIONS[idx]
        except Exception:
            return "Neutral"

    # ── Actions ───────────────────────────────────────────────
    def capture_mark(self):
        if not self.in_window:
            messagebox.showwarning("Outside Window",
                f"Attendance can only be marked between {WINDOW_START} and {WINDOW_END}.")
            return
        with self.frame_lock:
            frame = self.current_frame
        if frame is None:
            messagebox.showinfo("No Frame", "Start the camera first.")
            return
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(gray, 1.1, 4, minSize=(60, 60))
        marked = 0
        for (x, y, w, h) in faces:
            face_img = frame[y:y+h, x:x+w]
            match = self._identify(face_img)
            emotion = self._detect_emotion(face_img)
            if match:
                sid, name, dist = match
                conf = 1.0 - dist
                self.recorder.mark(sid, name, emotion, conf)
                self._add_row(sid, name, "Present", emotion,
                              datetime.datetime.now().strftime("%H:%M:%S"))
                marked += 1
        self._update_stats()
        self.status_bar.config(text=f"✅ Marked {marked} student(s) present")

    def register_student(self):
        sid  = self.reg_id.get().strip()
        name = self.reg_name.get().strip()
        if not sid or not name:
            messagebox.showwarning("Input Error", "Enter both Student ID and Name.")
            return
        with self.frame_lock:
            frame = self.current_frame
        if frame is None:
            messagebox.showinfo("No Frame", "Start the camera first.")
            return
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(gray, 1.1, 4, minSize=(60, 60))
        if len(faces) == 0:
            messagebox.showinfo("No Face", "No face detected. Position face in camera.")
            return
        (x, y, w, h) = faces[0]
        face_img = frame[y:y+h, x:x+w]
        emb = self._get_embedding(face_img)
        self.db.add_student(sid, name, emb)
        messagebox.showinfo("Registered", f"✅ Student '{name}' (ID: {sid}) registered.")
        self.status_bar.config(text=f"Registered: {name}")

    def mark_all_absent(self):
        self.recorder.mark_absents(self.db.students)
        for sid, rec in self.recorder.records.items():
            if rec["status"] == "Absent":
                exists = any(self.tree.item(i)["values"][0] == sid
                             for i in self.tree.get_children())
                if not exists:
                    self._add_row(sid, rec["name"], "Absent", "N/A", "N/A")
        self._update_stats()
        self.status_bar.config(text="Absents marked for undetected students")

    def _add_row(self, sid, name, status, emotion, t):
        tag = "present" if status == "Present" else "absent"
        self.tree.insert("", 0, values=(sid, name, status, emotion, t), tags=(tag,))

    def _update_stats(self):
        present = sum(1 for r in self.recorder.records.values() if r["status"] == "Present")
        absent  = sum(1 for r in self.recorder.records.values() if r["status"] == "Absent")
        total   = len(self.db.students)
        self.stat_present.config(text=str(present))
        self.stat_absent.config(text=str(absent))
        self.stat_total.config(text=str(total))

    def save_excel(self):
        path = self.recorder.save_excel()
        messagebox.showinfo("Saved", f"✅ Attendance saved to:\n{os.path.abspath(path)}")

    def save_csv(self):
        path = self.recorder.save_csv()
        messagebox.showinfo("Saved", f"✅ CSV saved to:\n{os.path.abspath(path)}")

    def clear_log(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        self.recorder.records.clear()
        self._update_stats()

    def on_close(self):
        self.stop_camera()
        self.root.destroy()


def main():
    root = tk.Tk()
    app = AttendanceApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
